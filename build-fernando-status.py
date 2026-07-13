#!/usr/bin/env python3
"""Regenerate the Fernando client status page DATA block from the audit spreadsheet.

Usage:
    python build-fernando-status.py [path-to-xlsx] [--date YYYY-MM-DD]

Reads the 23-row listing audit (migration + audit-status columns only — notes and
priority are deliberately NOT surfaced to the client) and rewrites the DATA block in
clients/fernando-3f9a2/status/index.html. Edit project PHASES below, then rerun.
"""
import sys, os, re, json, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PAGE = os.path.join(HERE, "clients", "fernando-3f9a2", "status", "index.html")
DEFAULT_XLSX = os.path.join(os.path.expanduser("~"), "Downloads", "Fernando-Listing-Audit.xlsx")
CLIENT = "Caribbean Dream Properties"

# ── EDIT HERE: project roadmap. status = done | active | scheduled | pending ─────────
PHASES = [
    dict(kind="core", when="Jul 2026", status="scheduled",
         en=dict(title="Platform migration",
                 desc="Move all 23 listings from Hostex to Hostaway — your new single control panel — within the first four weeks."),
         es=dict(title="Migración de plataforma",
                 desc="Trasladar las 23 unidades de Hostex a Hostaway — tu nuevo panel de control único — durante las primeras cuatro semanas.")),
    dict(kind="core", when="Jul–Aug 2026", status="scheduled",
         en=dict(title="Listing optimization",
                 desc="Review and correct every listing's title, description, photos, amenities and pricing accuracy."),
         es=dict(title="Optimización de anuncios",
                 desc="Revisar y corregir el título, la descripción, las fotos, los servicios y la exactitud de precios de cada anuncio.")),
    dict(kind="core", when="Ongoing", status="scheduled",
         en=dict(title="Airbnb & VRBO tuning",
                 desc="Optimize your existing Airbnb and VRBO presence for ranking and conversion."),
         es=dict(title="Optimización de Airbnb y VRBO",
                 desc="Optimizar tu presencia actual en Airbnb y VRBO para mejorar el posicionamiento y la conversión.")),
    dict(kind="channel", when="Aug–Sep 2026", status="scheduled",
         en=dict(title="Launch Booking.com",
                 desc="Connect and go live on Booking.com to reach new guest audiences."),
         es=dict(title="Lanzar Booking.com",
                 desc="Conectar y activar Booking.com para llegar a nuevos públicos de huéspedes.")),
    dict(kind="channel", when="Sep 2026", status="scheduled",
         en=dict(title="Launch Expedia",
                 desc="Add Expedia where a VRBO/Expedia connection is available in the Dominican Republic."),
         es=dict(title="Lanzar Expedia",
                 desc="Añadir Expedia donde exista una conexión VRBO/Expedia disponible en la República Dominicana.")),
    dict(kind="channel", when="Sep 2026", status="scheduled",
         en=dict(title="Launch Hopper",
                 desc="Connect Hopper to capture its mobile-first, younger booking audience."),
         es=dict(title="Lanzar Hopper",
                 desc="Conectar Hopper para captar su público de reservas móvil y más joven.")),
    dict(kind="core", when="Sep 2026", status="scheduled",
         en=dict(title="Dynamic pricing",
                 desc="Set up automated pricing that adjusts nightly rates to demand, events and season."),
         es=dict(title="Precios dinámicos",
                 desc="Configurar precios automáticos que ajustan las tarifas por noche según demanda, eventos y temporada.")),
    dict(kind="core", when="Aug 2026", status="scheduled",
         en=dict(title="Guest automations",
                 desc="Automate guest messaging, check-in instructions and review requests across every channel."),
         es=dict(title="Automatizaciones para huéspedes",
                 desc="Automatizar los mensajes a huéspedes, las instrucciones de check-in y las solicitudes de reseña en todos los canales.")),
    dict(kind="core", when="By Sep 2026", status="scheduled",
         en=dict(title="Website & direct booking engine",
                 desc="Rebuild your website with a new direct-booking engine so guests can book you commission-free — live by September."),
         es=dict(title="Sitio web y motor de reservas directas",
                 desc="Reconstruir tu sitio web con un nuevo motor de reservas directas para que los huéspedes reserven sin comisión de terceros — activo en septiembre.")),
    dict(kind="core", when="Ongoing", status="scheduled",
         en=dict(title="Biweekly strategy calls",
                 desc="Recurring calls to review performance, pricing and the next steps together."),
         es=dict(title="Llamadas quincenales de estrategia",
                 desc="Llamadas periódicas para revisar el rendimiento, los precios y los próximos pasos juntos.")),
]
# ─────────────────────────────────────────────────────────────────────────────────────

MIG = {"verified": "verified", "connected": "connected"}   # anything else -> "pending"
AUD = {"done": "done", "in progress": "progress"}           # anything else -> "pending"
FIRST, LAST = 9, 31  # data rows in the audit sheet

def read_listings(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Listing Audit"]
    out = []
    for i, r in enumerate(range(FIRST, LAST + 1), start=1):
        name = ws.cell(r, 2).value  # col B
        mig  = str(ws.cell(r, 4).value or "").strip().lower()   # col D
        aud  = str(ws.cell(r, 17).value or "").strip().lower()  # col Q
        out.append(dict(
            name=str(name).strip() if name else f"Unit {i}",
            migration=MIG.get(mig, "pending"),
            audit=AUD.get(aud, "pending"),
        ))
    return out

def main():
    args = [a for a in sys.argv[1:]]
    date = datetime.date.today().isoformat()
    if "--date" in args:
        date = args[args.index("--date") + 1]
        args = [a for a in args if a != "--date" and a != date]
    xlsx = args[0] if args else DEFAULT_XLSX

    listings = read_listings(xlsx) if os.path.exists(xlsx) else [
        dict(name=f"Unit {i}", migration="pending", audit="pending") for i in range(1, 24)
    ]
    if not os.path.exists(xlsx):
        print(f"! spreadsheet not found at {xlsx} — seeding 23 queued units")

    data = dict(
        updated=date, client=CLIENT,
        channelsTotal=sum(1 for p in PHASES if p["kind"] == "channel"),
        phases=PHASES, listings=listings,
    )
    block = "/* DATA-START — regenerated by build-fernando-status.py; do not hand-edit below */\nconst DATA = " \
        + json.dumps(data, ensure_ascii=False, indent=2) + ";\n/* DATA-END */"

    html = open(PAGE, encoding="utf-8").read()
    new, n = re.subn(r"/\* DATA-START.*?/\* DATA-END \*/", lambda m: block, html, flags=re.DOTALL)
    if n != 1:
        sys.exit(f"ERROR: expected exactly 1 DATA block, found {n}")
    open(PAGE, "w", encoding="utf-8").write(new)

    migrated = sum(1 for l in listings if l["migration"] == "verified")
    audited = sum(1 for l in listings if l["audit"] == "done")
    print(f"✓ {PAGE}")
    print(f"  {len(listings)} units · {migrated} migrated · {audited} audited · {len(PHASES)} phases · updated {date}")

if __name__ == "__main__":
    main()
